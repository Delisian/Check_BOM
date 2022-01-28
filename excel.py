import openpyxl
from openpyxl import load_workbook
import re
from PyQt5.QtWidgets import QTreeWidgetItem, QTreeView, QTreeWidget
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5 import QtCore
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors, Color


class ExcelFile:

    def __init__(self, path, text_browser):
        self.error_counter = 0
        self.fix = []
        self.words_regex = re.compile(r'[А-ЯЪЁ]+', re.I)
        self.blank_regex = re.compile(r'^\s+$')
        self.letter_regex = re.compile(r'^([A-Z]+)')
        self.string_regex = re.compile(r'([A-Z]\d{1,3})')
        self.designator_regex = re.compile(r'(([A-Z]+\d+(, )?)+)')
        self.textBrowser = text_browser
        self.workbook = load_workbook(path)
        self.sheets = self.workbook.sheetnames
        self.sheet = self.workbook[self.sheets[0]]
        self.title_check()
        self.date_check()
        self.device_name_check()
        self.blank_check()
        self.naming_check()
        self.designator_check()
        self.case_check()
        self.number_check()
        self.manufacturer_check()
        self.case_description_check()
        self.quantity_check()

    def title_check(self):
        if self.sheet["A1"].value != "Перечень элементов":                                                              # Project title check
            self.error_counter += 1
            QTreeWidgetItem(self.textBrowser, [f"{self.error_counter}) Ячейка A1 отведена под запись: "
                                               f"Перечень элементов на устройство"
                                               f"\nТекущее значение: {self.sheet['A1'].value}"])
                                               # f"\nИсправить?\n"])
        #     title1.setCheckState(1, QtCore.Qt.Unchecked)
        #     if designator_regex.fullmatch(self.sheet["A1"].value) is not None:
        #         self.fix.append()
        #         pass
        #     else:
        #         self.fix.append([title1, 'A', '1', "Перечень элементов на устройство"])
        name_regex = re.compile(r'((Плата)|(Модуль)) [A-Z]+[a-z]*\d*\.\d{4}\w?-?\d{,2}\w?')
        try:
            if name_regex.fullmatch(self.sheet["B1"].value) is None:
                self.error_counter += 1
                QTreeWidgetItem(self.textBrowser, [f"{self.error_counter}) Ячейка В1 предназначена для наименования "
                                                   f"проекта. "
                                                   f"\n текущее значение: {self.sheet['B1'].value}"
                                                   f"\nПроверьте корректность записи названия проекта"])
        except TypeError:
            self.error_counter += 1
            QTreeWidgetItem(self.textBrowser, [f"{self.error_counter}) Ячейка В1, предназначенная для наименования "
                                               f"проекта, пуста"])
            cell_range = self.sheet['C1':'E1']
            for row in cell_range:
                for cell in row:
                    try:
                        if cell.value is not None and cell.value != "" and \
                                self.blank_regex.search(cell.value) is None:
                            QTreeWidgetItem(self.textBrowser, [f"Ячейка {str(cell.column_letter)}{str(cell.row)} "
                                                               f"должна быть пуста"])
                    except TypeError:
                        pass

    def date_check(self):                                                                                               # Date format check
        date_regex = re.compile(r'^\d{2}\.\d{2}\.((\d{4})|(\d{2}))$')
        try:
            if date_regex.search(self.sheet["F1"].value) is None:
                self.error_counter += 1
                QTreeWidgetItem(self.textBrowser, [f"{self.error_counter}) Некорректный формат даты в ячейке F1: "
                                                   f"{self.sheet['F1'].value}"])
            else:
                pass
        except TypeError:
            self.error_counter += 1
            QTreeWidgetItem(self.textBrowser,
                            [f"{self.error_counter}) Ячейка F1, предназначенная для даты составления (правки) перечня, "
                             f"пуста"])

    def device_name_check(self):                                                                                        # Device naming format check
        device_regex = re.compile(r'Устройство: ("|«)([А-ЯЪЁа-яъё\w\-\d ]*?)("|»)')
        if self.sheet['A2'].value is not None and self.sheet['A2'].value != "" and \
           self.blank_regex.fullmatch(self.sheet['A2'].value) is None \
           and device_regex.fullmatch(self.sheet['A2'].value) is None:
            self.error_counter += 1
            QTreeWidgetItem(self.textBrowser, [f"{self.error_counter}) Некорректный формат записи в ячейке А2: "
                                               f"\n{self.sheet['A2'].value}"])

    def blank_check(self):                                                                                              # Blank lines format check
        cell_range1 = self.sheet['B2':'F2']
        for row in cell_range1:
            for cell in row:
                try:
                    if cell.value != "" and self.blank_regex.fullmatch(cell.value) is None:
                        self.error_counter += 1
                        QTreeWidgetItem(self.textBrowser, [f"{self.error_counter}) "
                                                           f"Ячейка {str(cell.column_letter)}{str(cell.row)} "
                                                           f"должна быть пуста"])
                except TypeError:
                    pass
        cell_range2 = self.sheet['A3':'F3']
        for row in cell_range2:
            for cell in row:
                try:
                    if cell.value != "" and self.blank_regex.fullmatch(cell.value) is None:
                        self.error_counter += 1
                        QTreeWidgetItem(self.textBrowser, [f"{self.error_counter}) "
                                                           f"Ячейка {str(cell.column_letter)}{str(cell.row)} "
                                                           f"должна быть пуста"])
                except TypeError:
                    pass
        cell_range3 = self.sheet['A5':'F5']
        for row in cell_range3:
            for cell in row:
                try:
                    if cell.value != "" and self.blank_regex.search(cell.value) is None:
                        self.error_counter += 1
                        QTreeWidgetItem(self.textBrowser, [f"{self.error_counter}) "
                                                           f"Ячейка {str(cell.column_letter)}{str(cell.row)} "
                                                           f"должна быть пуста"])
                except TypeError:
                    pass

    def naming_check(self):                                                                                             # Column names check
        cell_range3 = self.sheet['A4':'F4']
        correct = [["Позиционное обозначение", "Позиц. обозн.", "Поз. об."], ["Наименование", "Наимен.", "Наим."],
                   ["Аналог"], ["Корпус", "Корп.", "Кор."], ["Количество", "Кол-во", "Кол."],
                   ["Примечание", "Примеч.", "Прим."]]
        suggested = ["Позиционное обозначение", "Наименование", "Аналог", "Корпус", "Кол-во", "Прим."]
        counter = 0
        for cell in cell_range3[0]:
            try:
                if cell.value not in correct[counter]:
                    self.error_counter += 1
                    QTreeWidgetItem(self.textBrowser, [f"{self.error_counter}) "
                                                       f"Ячейка {str(cell.column_letter)}{str(cell.row)}"
                                                       f" имеет текущее значение: {str(cell.value)}"
                                                       f"\nПредполагаемое значение: {str(suggested[counter])}"])
            except AttributeError:
                self.error_counter += 1
                QTreeWidgetItem(self.textBrowser, [f"{self.error_counter}) "
                                                   f"Ячейка {str(cell.column_letter)}{str(cell.row)}"
                                                   f" пуста\nПредполагаемое значение: {str(suggested[counter])}"])
            counter += 1

    def designator_check(self):                                                                                         # Designator format check
        for cell in self.sheet["A"]:
            if cell.row > 5:
                try:
                    if self.designator_regex.fullmatch(cell.value) is None:
                        self.error_counter += 1
                        QTreeWidgetItem(self.textBrowser, [f"{self.error_counter}) "
                                                           f"Некорректный формат записи указателя в ячейке A"
                                                           f"{str(cell.row)}: {str(cell.value)}"])
                except TypeError:
                    pass
        for cell in self.sheet["A"]:
            if cell.row > 5:
                try:                                                                                                    #Blank line check
                    if self.designator_regex.fullmatch(cell.value) is not None:
                        if self.sheet["A"].index(cell) + 1 < len(self.sheet["A"]):
                            if str(self.letter_regex.search(self.sheet["A"][(self.sheet["A"].index(cell) + 1)].value)) \
                                    != str(self.letter_regex.search(cell.value)) and \
                                    self.letter_regex.search(self.sheet["A"][(self.sheet["A"].index(cell) + 1)].value) \
                                    is not None:
                                self.error_counter += 1
                                QTreeWidgetItem(self.textBrowser, [f"{self.error_counter}) Отсутствует отступ между "
                                                                   f"разными типами компонентов в строках "
                                                                   f"{str(cell.row)} и {str(int(cell.row) + 1)}"])
                except TypeError:
                    pass

    def case_check(self):                                                                                               # Case format check
        case_check_regex = re.compile(r'^(\w\d+-?\w?)+$')                                                               # Nonstandart case check
        CR_regex = re.compile(r'[CLR]')                                                                                 # C, L or R avaibility check
        CR_case_regex = re.compile(r'^\w\d{4}$|^\d{4}$')                                                                # Standart capacitor's, inductor's and resistor's case check
        CR_array = []                                                                                                   # Array of rows that include capacitors, inductors or resistors
        for cell_a in self.sheet["A"]:                                                                                  # Case check for resistors, inductors and capacitors
            try:
                if str(self.letter_regex.search(cell_a.value)) == str(CR_regex.search(cell_a.value)) \
                        and self.letter_regex.search(cell_a.value) is not None:
                    if CR_regex.search(cell_a.value) is not None:
                        CR_array.append(cell_a.row)
            except TypeError:
                pass
        for cell in self.sheet["D"]:
            if cell.row > 5:
                try:
                    if cell.value != "" and self.blank_regex.fullmatch(cell.value) is None:
                        if cell.row in CR_array:
                            if CR_case_regex.fullmatch(cell.value) is None:
                                self.error_counter += 1
                                QTreeWidgetItem(self.textBrowser, [f"{self.error_counter}) Некорректный формат записи "
                                                                   f"корпуса в ячейке D{str(cell.row)}: "
                                                                   f"{str(cell.value)}"])
                        else:
                            if case_check_regex.fullmatch(cell.value) is None:
                                self.error_counter += 1
                                QTreeWidgetItem(self.textBrowser, [f" {self.error_counter}) Некорректный формат записи "
                                                                   f"корпуса в ячейке D{str(cell.row)}: "
                                                                   f"{str(cell.value)}"])
                            else:
                                pass
                    else:
                        self.error_counter += 1
                        QTreeWidgetItem(self.textBrowser, [f"{self.error_counter}) В ячейке "
                                                           f"{str(cell.column_letter)}{str(cell.row)} отсутствует "
                                                           f"запись корпуса компонента"])

                except TypeError:
                    pass

    def number_check(self):
        self.num_check = []                                                                                             # Number format check
        number_check_regex = re.compile(r'^\d+$')
        for cell in self.sheet["E"]:
            try:
                if self.words_regex.fullmatch(cell.value) is None and str(cell.row) != "3":
                    if number_check_regex.fullmatch(cell.value) is None:
                        self.error_counter += 1
                        QTreeWidgetItem(self.textBrowser, [f"{self.error_counter}) Некорректный формат записи количества"
                                                           f" компонентов в ячейке: "
                                                           f"{str(cell.column_letter)}{str(cell.row)}: "
                                                           f"{str(cell.value)}"])
                else:
                    self.num_check.append(cell)                                                                         # Correct numbers
            except TypeError:
                # print("\nBlank line")
                pass

    def manufacturer_check(self):                                                                                       # Manufacturer format check
        manufacturer_check_regex = re.compile(r'([A-Z]+ ?[a-z]*)+\.?')
        for cell in self.sheet["F"]:
            try:
                if cell.value is self.sheet["F1"].value:
                    pass
                else:
                    if self.blank_regex.search(cell.value) is None and cell.value != "" and cell.value != "'":
                        if self.words_regex.search(cell.value) is None:
                            if manufacturer_check_regex.match(cell.value) is None:
                                self.error_counter += 1
                                QTreeWidgetItem(self.textBrowser, [f"{self.error_counter}) Проверьте корректность "
                                                                   f"записи названия компании в ячейке "
                                                                   f"F{str(cell.row)}. Текущее значение: "
                                                                   f"{str(cell.value)}"])
                    else:
                        pass
            except TypeError:
                # print("\nBlank line")
                pass

    def case_description_check(self):                                                                                   # Condensator's, inductors and resistor's case-description ->
        case_regex = re.compile(r'^\d{4}|\d{4}$')                                                                       # -> column(B and D) comparison
        CR_regex = re.compile(r'[CLR]')
        for cell in self.sheet["A"]:
            try:
                if str(self.letter_regex.search(cell.value)) == str(CR_regex.search(cell.value)) \
                        and self.letter_regex.search(cell.value) is not None:
                    current_index = self.sheet["A"].index(cell)
                    if str(case_regex.findall(self.sheet["B"][current_index].value)) \
                            != str(case_regex.findall(self.sheet["D"][current_index].value)):
                        self.error_counter += 1
                        QTreeWidgetItem(self.textBrowser, [f"{self.error_counter}) Обозначения корпуса в ячейках "
                                                           f"B{cell.row} ({self.sheet['B'][current_index].value})"
                                                           f" и D{cell.row} ({self.sheet['D'][current_index].value}) "
                                                           f"не совпадают"])
            except TypeError:
                pass

    def quantity_check(self):                                                                                           # Quantity check
        item_regex = re.compile(r'([A-Z]+\d+)')
        for cell in self.sheet["A"]:
            try:
                if cell in self.num_check:
                    current_index = self.sheet["A"].index(cell)
                    matches = item_regex.findall(cell)
                    if str(len(matches)) != str(self.sheet["E"][current_index].value):
                        self.error_counter += 1
                        QTreeWidgetItem(self.textBrowser,
                                        [f"{self.error_counter}) Количество компонентов в колонке указателей и колонке "
                                         f"\"Количество\" не совпадают в строке {cell.row} "
                                         f"\n текущие значения: {cell.value}, "
                                         f"{self.sheet['E'][current_index].value}"])
            except TypeError:
                pass

    def save(self, filename):
        self.workbook.save(filename=filename)
    # def checkbox_fix(self):
    #     for checkbox in self.fix:
    #         if checkbox[0].checkState(1) == QtCore.Qt.Checked:
    #             if checkbox[3] == "Перечень элементов на устройство":




# if __name__ == '__main__':
#     file = ExcelFile(path="FP-APC61850.1021A.xlsx")